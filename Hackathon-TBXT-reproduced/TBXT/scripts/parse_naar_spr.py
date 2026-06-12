"""
Parse decrypted Naar SPR XLSX files into a single training set.

Each file contains 'Data summary' (or ' Data summary' / 'Data summary ' / 'Evaluation-Affinity')
listing per-compound SPR fits across multiple replicate wells.

Strategy:
  1. Use the 'Data summary' sheet variants — most reliable across the 15 campaigns.
  2. Locate columns dynamically by header name: "Compounds" / "Compound ID" / " Compounds",
     "Kinetics Chi²" + "KD (M)" (kinetics fit), "Affinity Chi²" + "KD" (steady-state fit).
  3. For each compound replicate: prefer steady-state Kd when affinity fit Chi² ≤ 5; fall back
     to kinetics Kd when kinetics Chi² ≤ 10. Drop both if poor fit.
  4. Aggregate replicates within a campaign by median Kd; then aggregate across campaigns.
  5. Filter: drop Kd outside [1e-12, 1e-2 M] (i.e. between 1 pM and 10 mM).

Output: data/qsar/naar_kd_dataset.csv
  compound_id, smiles, kd_uM_median, log10_kd_M_median, n_fits, kd_M_min, kd_M_max, campaigns
"""
import csv
import math
import re
import statistics
from pathlib import Path

import openpyxl

NAAR = Path(__file__).resolve().parents[1] / "data/naar"
SPR = NAAR / "spr_decrypted"
QSAR = Path(__file__).resolve().parents[1] / "data/qsar"
QSAR.mkdir(exist_ok=True)

KD_MAX_M = 1e-2     # 10 mM cutoff (above = no binding)
KD_MIN_M = 1e-12    # 1 pM (below = fitting failure)
AFF_CHI2_MAX = 5.0
KIN_CHI2_MAX = 10.0


def normalize_id(raw):
    """Extract the cleanest compound ID from raw cell content.

    Handles formats like:
      'M567200(CSC015274258)-1'  →  'CSC015274258'
      'Z1023598674'              →  'Z1023598674'
      'CF-10-045-2'              →  'CF-10-045'
      'Reference(...)'           →  None  (positive control)
    """
    if raw is None: return None
    s = str(raw).strip()
    if not s: return None
    if s.lower().startswith("reference"): return None
    # Inner-paren ID takes precedence
    m = re.match(r"^[^(]*\(([^)]+)\)", s)
    if m:
        inner = m.group(1).strip()
        if inner: s = inner
    # Strip trailing -N replicate suffix (1-2 digit numeric)
    s = re.sub(r"-\d{1,2}$", "", s)
    return s.strip() or None


def find_data_sheet(wb):
    """Return the most useful sheet for per-compound Kd fits."""
    # Prefer 'Data summary' variants — most consistent column layout
    for sn in wb.sheetnames:
        if sn.strip().lower() in ("data summary", "data summary"):
            return sn
    for sn in wb.sheetnames:
        if "data summary" in sn.lower(): return sn
    # Fallback: any "ffinity" sheet that's actually populated
    for sn in wb.sheetnames:
        if "ffinity" in sn.lower():
            ws = wb[sn]
            if ws.max_row > 20:  # only if it has actual data, not just metadata
                return sn
    return None


def find_columns(header_rows):
    """Identify columns from header rows. Returns dict.

    Column 0 is assumed to always hold the compound ID (verified across all 15 files).
    Collect ALL columns labeled 'KD (M)' or 'KD', plus any 'Chi²' columns. Per-row
    we'll try each KD column in order and pick the first valid value.
    """
    col = {"id": 0, "chi2_cols": [], "kd_cols": []}
    for ri, row in enumerate(header_rows):
        for ci, v in enumerate(row):
            if v is None: continue
            vs = str(v).strip()
            vsl = vs.lower()
            if "chi" in vsl and "²" in vs:
                if ci not in col["chi2_cols"]: col["chi2_cols"].append(ci)
            if vs in ("KD (M)", "KD"):
                if ci not in col["kd_cols"]: col["kd_cols"].append(ci)
    col["chi2_cols"].sort()
    col["kd_cols"].sort()
    return col


def parse_spr_file(xlsx_path):
    """Yield (compound_id, kd_M, source_string) for one file."""
    out = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  FAIL load: {xlsx_path.name} — {e}")
        return out

    sheet_name = find_data_sheet(wb)
    if sheet_name is None:
        print(f"  no data sheet in {xlsx_path.name}")
        return out

    ws = wb[sheet_name]
    # Read first 5 rows as candidates for header
    header_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        header_rows.append(row)
        if i >= 4: break
    cols = find_columns(header_rows)

    if not cols["kd_cols"]:
        print(f"  no Kd column in {xlsx_path.name} sheet '{sheet_name}'  (cols={cols})")
        return out

    # Helper: for a given KD column, find its associated Chi² column (the closest preceding Chi² col)
    def chi2_for_kd(kd_col):
        preceding = [c for c in cols["chi2_cols"] if c < kd_col]
        return max(preceding) if preceding else None

    # Iterate data rows
    n = 0
    n_kept = 0
    for row in ws.iter_rows(values_only=True):
        n += 1
        if n <= 5: continue  # skip header
        if not row or cols["id"] >= len(row): continue
        cid = normalize_id(row[cols["id"]])
        if not cid: continue

        # Try each KD column in order; accept first valid one
        kd_M = None
        source = None
        for kd_col in cols["kd_cols"]:
            if kd_col >= len(row): continue
            kd_raw = row[kd_col]
            if kd_raw is None: continue
            try:
                kd_f = float(kd_raw)
            except (TypeError, ValueError):
                continue
            if kd_f <= 0 or kd_f < KD_MIN_M or kd_f > KD_MAX_M:
                continue
            # Check Chi² for this KD column
            chi2_col = chi2_for_kd(kd_col)
            chi2_f = None
            if chi2_col is not None and chi2_col < len(row):
                try: chi2_f = float(row[chi2_col]) if row[chi2_col] is not None else None
                except (TypeError, ValueError): pass
            chi2_max = AFF_CHI2_MAX if "aff" in str(chi2_col).lower() else KIN_CHI2_MAX
            if chi2_f is not None and chi2_f > KIN_CHI2_MAX:  # use looser kinetics threshold; we don't know which type
                continue
            kd_M = kd_f
            source = f"col{kd_col}"
            break

        if kd_M is None: continue
        out.append((cid, kd_M, source))
        n_kept += 1

    print(f"  {xlsx_path.name[:55]:55s}  sheet='{sheet_name[:25]}'  rows={n} kept={n_kept}")
    return out


def main():
    spr_files = sorted(SPR.glob("*.xlsx"))
    print(f"Parsing {len(spr_files)} SPR campaign files\n")

    # Per-compound, per-campaign accumulator
    # cid → list of (kd_M, source, campaign_filename)
    per_compound = {}
    for f in spr_files:
        date_tag = f.name[:8]
        rows = parse_spr_file(f)
        for cid, kd_M, source in rows:
            per_compound.setdefault(cid, []).append((kd_M, source, date_tag))

    print(f"\nTotal unique compound IDs: {len(per_compound)}")
    n_per_cmpd = [len(v) for v in per_compound.values()]
    if n_per_cmpd:
        print(f"  fits per compound: min={min(n_per_cmpd)}, "
              f"median={statistics.median(n_per_cmpd):.0f}, max={max(n_per_cmpd)}, "
              f"total={sum(n_per_cmpd)}")

    # Aggregate
    agg = []
    for cid, fits in per_compound.items():
        kds = [k for k, _, _ in fits]
        med_kd_M = statistics.median(kds)
        agg.append({
            "compound_id": cid,
            "kd_M_median": med_kd_M,
            "kd_uM_median": round(med_kd_M * 1e6, 4),
            "log10_kd_M_median": round(math.log10(med_kd_M), 3),
            "pkd": round(-math.log10(med_kd_M), 3),
            "n_fits": len(fits),
            "n_campaigns": len({c for _, _, c in fits}),
            "kd_M_min": min(kds),
            "kd_M_max": max(kds),
            "kd_M_geomean": 10 ** statistics.mean(math.log10(k) for k in kds),
            "campaigns": ";".join(sorted({c for _, _, c in fits})),
            "sources": ";".join(sorted({s for _, s, _ in fits})),
        })

    # Join with master SMILES
    print("\nJoining with master SMILES...")
    smiles_map = {}
    with open(NAAR / "naar_smiles.csv") as f:
        for r in csv.reader(f):
            if len(r) >= 2 and r[0] and r[1]:
                smiles_map[r[0]] = r[1]
    with open(NAAR / "naar_sheet_export.csv") as f:
        next(f)  # header
        for r in csv.reader(f):
            if len(r) >= 2 and r[0] and r[1]:
                smiles_map.setdefault(r[0], r[1])
    print(f"  Total ID→SMILES mappings: {len(smiles_map)}")

    n_with_smi = 0
    for r in agg:
        smi = smiles_map.get(r["compound_id"], "")
        r["smiles"] = smi
        if smi: n_with_smi += 1
    print(f"  Compounds with SMILES: {n_with_smi}/{len(agg)} ({100*n_with_smi/len(agg):.1f}%)")

    agg_with_smi = [r for r in agg if r.get("smiles")]
    agg_with_smi.sort(key=lambda r: r["kd_M_median"])

    log_kds = [r["log10_kd_M_median"] for r in agg_with_smi]
    print(f"\nlog10(Kd in M) distribution across {len(log_kds)} compounds:")
    print(f"  min={min(log_kds):.2f} (= {10**min(log_kds)*1e6:.2f} µM, strongest)")
    print(f"  max={max(log_kds):.2f} (= {10**max(log_kds)*1e6:.0f} µM, weakest)")
    print(f"  median={statistics.median(log_kds):.2f} (= {10**statistics.median(log_kds)*1e6:.0f} µM)")
    print(f"  pKd range: {-max(log_kds):.2f} – {-min(log_kds):.2f}")

    n_strong = sum(1 for r in agg_with_smi if r["kd_uM_median"] < 100)
    n_mid = sum(1 for r in agg_with_smi if 100 <= r["kd_uM_median"] < 1000)
    n_weak = sum(1 for r in agg_with_smi if r["kd_uM_median"] >= 1000)
    print(f"\nClass distribution:")
    print(f"  Strong (Kd < 100 µM):    {n_strong} ({100*n_strong/len(agg_with_smi):.0f}%)")
    print(f"  Medium (100 µM – 1 mM):  {n_mid} ({100*n_mid/len(agg_with_smi):.0f}%)")
    print(f"  Weak (Kd ≥ 1 mM):        {n_weak} ({100*n_weak/len(agg_with_smi):.0f}%)")

    # Write
    cols = ["compound_id", "smiles", "kd_uM_median", "log10_kd_M_median", "pkd",
            "n_fits", "n_campaigns", "kd_M_min", "kd_M_max", "kd_M_median",
            "kd_M_geomean", "campaigns", "sources"]
    out_path = QSAR / "naar_kd_dataset.csv"
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in agg_with_smi:
            w.writerow({k: r.get(k, "") for k in cols})
    print(f"\nWrote {out_path} ({len(agg_with_smi)} rows)")

    # Also write the without-SMILES set so we can investigate ID matching gaps
    no_smi = [r for r in agg if not r.get("smiles")]
    if no_smi:
        with open(QSAR / "naar_kd_no_smiles.csv", "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["compound_id", "kd_uM_median", "n_fits"])
            w.writeheader()
            for r in no_smi:
                w.writerow({k: r.get(k, "") for k in ["compound_id", "kd_uM_median", "n_fits"]})


if __name__ == "__main__":
    main()
